from io import BytesIO
from zipfile import ZipFile

from openpyxl import load_workbook
from pypdf import PdfReader
from reportlab.pdfgen.canvas import Canvas

from evidence_pdf.extractor import extract_evidence
from evidence_pdf.exporters import build_export_bundle
from evidence_pdf.parser import read_pages
from evidence_pdf.utils import infer_meta, split_terms


def make_pdf() -> bytes:
    output = BytesIO()
    canvas = Canvas(output)
    canvas.drawString(72, 760, "ACME 2024 Annual Report")
    canvas.showPage()
    canvas.drawString(72, 760, "R&D expenditure was USD 12.50 million in 2024.")
    canvas.showPage()
    canvas.save()
    return output.getvalue()


def test_end_to_end_pipeline():
    pdf = make_pdf()
    pages = read_pages(pdf)
    meta = infer_meta("ACME_2024_Annual_Report.pdf", pages)
    results = extract_evidence(pages, "研发投入", meta)

    assert len(results) == 1
    assert results[0].page_number == 2
    assert "12.50" in results[0].value_candidate
    assert results[0].evidence_filename.endswith("_p2.pdf")

    files = build_export_bundle(results, {meta.source_filename: pdf})
    evidence_pdf = files[f"证据页/{results[0].evidence_filename}"]
    assert len(PdfReader(BytesIO(evidence_pdf)).pages) == 1

    workbook = load_workbook(BytesIO(files["提取结果.xlsx"]))
    assert workbook.active.max_row == 2
    assert workbook.active["I2"].value == 2

    with ZipFile(BytesIO(files["证据提取结果包.zip"])) as archive:
        names = archive.namelist()
        assert "提取结果.xlsx" in names
        assert any(name.startswith("证据页/") for name in names)


def test_empty_and_scanned_like_pages_do_not_match():
    pdf = make_pdf()
    pages = read_pages(pdf)
    results = extract_evidence(pages, "not present", infer_meta("ACME_2024.pdf", pages))
    assert results == []


def test_content_metadata_and_false_positive_rejection():
    from evidence_pdf.models import PageText

    pages = [
        PageText(1, "宁德时代新能源科技股份有限公司\n2025年年度报告\n2026年3月"),
        PageText(20, "持续高强度的研发投入。公司拥有及申请的国内外专利总数达 54,538 项。"),
        PageText(30, "(3)近三年公司研发投入金额及占营业收入的比例\n"
                     "项目 2025年 2024年 2023年\n"
                     "研发投入金额(千元) 22,146,581 18,606,756 18,356,108"),
    ]
    meta = infer_meta("20260310105829_random.pdf", pages)
    results = extract_evidence(pages, "研发投入", meta)

    assert meta.company == "宁德时代新能源科技股份有限公司"
    assert meta.year == "2025"
    assert meta.language == "中文"
    assert len(results) == 1
    assert results[0].page_number == 30
    assert results[0].value_candidate == "22,146,581"
    assert results[0].amount_unit == "千元"
    assert results[0].currency == "原文未注明"
